"""XLSX export — Notes, References, Tags sheets."""

from __future__ import annotations

import io

from openpyxl import Workbook

from kit.engine.junction import get_tag_names_for_note
from kit.engine.sql import q
from kit.engine.runtime import Runtime


def export_xlsx(runtime: Runtime) -> bytes:
    conn = runtime.conn
    wb = Workbook()
    ws_notes = wb.active
    ws_notes.title = "Notes"

    ws_notes.append(["notebook_id", "id", "title", "body", "references", "status", "tags"])
    for row in conn.execute("SELECT * FROM notes ORDER BY notebook_id, id"):
        r = dict(row)
        body = (r.get("body") or "").replace("\x1e", " | ")
        tags = ", ".join(get_tag_names_for_note(conn, r["notebook_id"], r["id"]))
        ws_notes.append([
            r["notebook_id"], r["id"], r["title"], body,
            r.get("references") or "", r.get("status") or "", tags,
        ])

    ws_refs = wb.create_sheet("References")
    ws_refs.append(["id", "title", "link", "type", "summary", "created_at"])
    for row in conn.execute(f'SELECT * FROM {q("references")} ORDER BY title'):
        r = dict(row)
        ws_refs.append([r["id"], r["title"], r["link"], r.get("type") or "", r.get("summary") or "", r.get("created_at") or ""])

    ws_tags = wb.create_sheet("Tags")
    ws_tags.append(["id", "name", "description"])
    for row in conn.execute("SELECT * FROM tags ORDER BY name"):
        r = dict(row)
        ws_tags.append([r["id"], r["name"], r.get("description") or ""])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
